import openpyxl
import sys

wb = openpyxl.load_workbook('/Users/alaragebes/Desktop/solar_data/STATES/Louisiana.xlsx')
account = openpyxl.load_workbook('/Users/alaragebes/Desktop/solar_data/CS_Account&ContactList__2018-07-09 (1).xlsx')
customer = openpyxl.load_workbook('/Users/alaragebes/Desktop/solar_data/Shipments for Kerim 2016-2018 copy.xlsx')
zoho = openpyxl.load_workbook('/Users/alaragebes/Desktop/solar_data/Zoho Customer List (Accounts that Ordered).xlsx')

print wb.sheetnames
web = wb['Sheet1']
r = web.max_row
print r

print zoho.sheetnames
sheet = zoho['Zoho']
sheet_row = sheet.max_row

print customer.sheetnames
customers_2016 = customer['2016']
customers_2017 = customer['2017']
customers_2018 = customer['2018']
c_16 = customers_2016.max_row
c_17 = customers_2017.max_row
c_18 = customers_2018.max_row

accounts = account['account_list_for_yelp_data-a_ac']
rb = accounts.max_row
sample = 'Cross-Ref'



print "Number of solar installers on Yelp...%r" %(r-1)


def change_name(input):
    new_value = input.replace(' ','').decode('utf-8')
    short_value = new_value[:12]
    return short_value

## add short names to new sheet
def add_names_to_file (file, sheetname, rows, col, lists):
    list_names = list(lists)
    if sheetname in file.sheetnames:
        print "Sheetname already exists"
        sample = file[sheetname]
        if col == 3:
            sample.cell(row=1, column=col).value = "web and account"
        if col == 4:
            sample.cell(row=1, column=col).value = "Account"
        if col == 5:
            sample.cell(row=1, column=col).value = "web and customer"
        if col == 6:
            sample.cell(row=1, column=col).value = "Customer"
        for i in range(2, rows+2):
            sample.cell(row=i, column=col).value = list_names[i-2]
        file.save('/Users/alaragebes/Desktop/solar_data/STATES/Louisiana.xlsx')
    else:
        print "Creating new sheet %s" %(sheetname)
        file.create_sheet(sheetname)
        sample = file[sheetname]
        sample.cell(row=1, column=col).value = "Yelp"
        for i in range(2, rows+2):
            sample.cell(row=i, column=col).value = list_names[i-2]
        file.save('/Users/alaragebes/Desktop/solar_data/STATES/Louisiana.xlsx')
    print "Changes have been saved :)"



names = set()
for i in range(2, r+1):
    short_name = change_name((web.cell(row=i, column=1).value).encode('utf-8'))
    names.add(short_name)
print names
ru = len(names)
print "Number of company names: %r" %(len(names))
add_names_to_file(wb, sample, ru, 1, names)
print "************************************************************"


account_name = set()
for i in range(1, rb+1):
    if accounts.cell(row=i, column=4).value == "Louisiana":
        changed_name = unicode(accounts.cell(row=i, column=1).value).encode('utf-8')
        name_changed = change_name(changed_name)
        account_name.add(name_changed)
print account_name
num = len(account_name)
print "Number of account names: %r" %(num)
add_names_to_file(wb, sample, num, 4, account_name)
print "************************************************************"


customer_name = set()
for i in range(1, c_16+1):
    if (customers_2016.cell(row=i, column=13).value == 'Louisiana') or (customers_2016.cell(row=i, column=13).value == 'LA'):
        changed_name = change_name(unicode(customers_2016.cell(row=i, column=8).value).encode('utf-8'))
        customer_name.add(changed_name)

for i in range(1, c_17+1):
    if (customers_2017.cell(row=i, column=13).value == 'Louisiana') or (customers_2017.cell(row=i, column=13).value == 'LA'):
        changed_name = change_name(unicode(customers_2017.cell(row=i, column=8).value).encode('utf-8'))
        customer_name.add(changed_name)

for i in range(1, c_18+1):
    if (customers_2018.cell(row=i, column=13).value == 'Louisiana') or (customers_2018.cell(row=i, column=13).value == 'LA'):
        changed_name = change_name(unicode(customers_2018.cell(row=i, column=8).value).encode('utf-8'))
        customer_name.add(changed_name)

for i in range(1, sheet_row+1):
    if (sheet.cell(row=i, column=6).value == 'Louisiana') or (sheet.cell(row=i, column=6).value == 'LA'):
        changed_name = change_name(unicode(sheet.cell(row=i, column=2).value).encode('utf-8'))
        customer_name.add(changed_name)
print customer_name
number = len(customer_name)
print "Number of customer names: %r" %(number)
add_names_to_file(wb, sample, number, 6, customer_name)
print "************************************************************"

web_and_account = names & account_name
overlap = len(web_and_account)
print "The overlap between web and account: %r" %(overlap)
add_names_to_file(wb, sample, overlap, 3, web_and_account)
print "************************************************************"


web_and_customer = names & customer_name
overlap2 = len(web_and_customer)
print "The overlap between web and customer: %r" %(overlap2)
add_names_to_file(wb, sample, overlap2, 5, web_and_customer)
print "************************************************************"
