import openpyxl
import sys

##load file
wb = openpyxl.load_workbook('###')
change = wb['Cross-Ref']
sheet = wb['Sheet1']
ru = sheet.max_row
r = change.max_row

def change_name(input):
    new_value = input.replace(' ','')
    short_value = new_value[:12]
    return short_value

##add customer and account in list
common = []
for i in range(2,r+1):
    common.append(change.cell(row=i, column=3).value)

for i in range(2,r+1):
    common.append(change.cell(row=i, column=5).value)

##remove duplicates
result = set()
for name in common:
    if name != None :
        result.add(name)

print result
print len(result)

## get names
yelp = set()
for i in range(2, r+1):
    if change.cell(row=i, column=1).value != None:
        yelp.add(change.cell(row=i, column=1).value)

print len(yelp)

##remove the duplicates
sub = yelp - result
call_list = list(sub)
print "the number of new companies: %r" %(len(call_list))
change.cell(row=1, column=8).value = 'Call List'
change.cell(row=1, column=9).value = 'URL'
change.cell(row=1, column=10).value = 'Phone'
wb.save('###')



for i in range(2, (len(call_list)+2)):
    change.cell(row=i, column=8).value = call_list[i-2]
    for num in range(2, ru+1):
        if change_name(sheet.cell(row=num, column=1).value) == call_list[i-2]:
            change.cell(row=i, column=9).value = sheet.cell(row=num, column=6).value
            change.cell(row=i, column=10).value = sheet.cell(row=num, column=7).value
            wb.save('###')
